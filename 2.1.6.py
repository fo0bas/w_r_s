import sys
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QLabel, QVBoxLayout, QFileDialog, QTabWidget, QTableWidget, QTableWidgetItem, QSizePolicy, QLineEdit, QHBoxLayout, QComboBox, QMenu, QCalendarWidget
from PyQt5.QtGui import QFont, QColor, QIcon
import pandas as pd
from PyQt5.QtCore import QTimer
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
class MainWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Удержание Рейтинг Скорость v1.0.0")
        self.selected_file = None
        desktop = QApplication.desktop()
        screen_rect = desktop.screenGeometry()
        self.setGeometry(screen_rect)
        self.showMaximized()
        

        self.button = QPushButton("Выбрать файл", self)
        self.button.setStyleSheet("background-color: #ADFF2F; color: black; font-size: 20px;")
        self.button.clicked.connect(self.select_file)

        self.label = QLabel("", self)
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setStyleSheet("font-size: 20px;")

        self.tab_widget = QTabWidget(self)
        self.base_tab = QWidget()
        self.retention_tab = QWidget()
        self.pvz_tab = QWidget()  # Добавляем новую вкладку "Рейтинг ПВЗ"
        self.df = None
        self.tab_widget.addTab(self.retention_tab, "Удержание")
        self.tab_widget.addTab(self.base_tab, "База и поиск")
        self.tab_widget.addTab(self.pvz_tab, "Рейтинг ПВЗ")  # Добавляем вкладку "Рейтинг ПВЗ"

        self.tab_widget.setTabText(0, "Удержание")

        font = QFont()
        font.setPointSize(12)
        font.setBold(True)
        self.tab_widget.tabBar().setFont(font)

        self.table_widget = QTableWidget(self.base_tab)
        self.table_widget.setColumnCount(0)
        self.table_widget.setRowCount(0)

        self.base_layout = QVBoxLayout(self.base_tab)

        self.search_layout = QHBoxLayout()
        self.search_box = QLineEdit()
        self.city_filter = QComboBox()
        self.city_filter.addItems(["Выбрать город", "Екатеринбург", "Уфа", "Челябинск", "Магнитогорск"])
        self.search_button = QPushButton("Найти")
        self.search_reset_button = QPushButton("Сбросить и обновить")
        self.search_reset_button.setEnabled(False)
        self.search_layout.addWidget(self.search_box)
        self.search_layout.addWidget(self.city_filter)
        self.search_layout.addWidget(self.search_button)
        self.search_layout.addWidget(self.search_reset_button)
        self.base_layout.addLayout(self.search_layout)
        self.search_box.returnPressed.connect(self.search_data)

        self.base_layout.addWidget(self.table_widget)
        self.base_tab.setLayout(self.base_layout)

        # Добавление полей на вкладку "Удержание"
        retention_layout = QVBoxLayout(self.retention_tab)
        retention_layout.setSpacing(1)  # Установка расстояния между строками

        city_label = QLabel("Город:", self.retention_tab)
        self.city_combo = QComboBox(self.retention_tab)
        self.city_combo.addItems(["Выбрать город", "Екатеринбург", "Уфа", "Челябинск", "Магнитогорск"])
        city_label.setStyleSheet("margin-right: -10px;")
        retention_layout.addWidget(city_label)
        retention_layout.addWidget(self.city_combo)

        self.address_label = QLabel("Адрес:", self.retention_tab)
        self.address_combo = QComboBox(self.retention_tab)
        retention_layout.addWidget(self.address_label)
        retention_layout.addWidget(self.address_combo)

        self.city_combo.currentIndexChanged.connect(self.update_address_combo)

        barcode_label = QLabel("ШК:", self.retention_tab)
        self.barcode_edit = QLineEdit(self.retention_tab)
        retention_layout.addWidget(barcode_label)
        retention_layout.addWidget(self.barcode_edit)

        status_label = QLabel("Статус удержания:", self.retention_tab)
        self.status_combo = QComboBox(self.retention_tab)
        self.status_combo.addItems(["К удержанию", "В работе", "Удержано", "Доплачено"])
        retention_layout.addWidget(status_label)
        retention_layout.addWidget(self.status_combo)

        amount_label = QLabel("Сумма:", self.retention_tab)
        self.amount_edit = QLineEdit(self.retention_tab)
        retention_layout.addWidget(amount_label)
        retention_layout.addWidget(self.amount_edit)

        date_label = QLabel("Дата:", self.retention_tab)
        self.date_calendar = QCalendarWidget(self.retention_tab)
        self.date_calendar.setFixedSize(200, 200)  # Установка размеров 200x200 пикселей
        self.date_calendar.setSelectedDate(QDate(2024, 1, 1))
        retention_layout.addWidget(date_label)
        retention_layout.addWidget(self.date_calendar)

        name_label = QLabel("ФИО:", self.retention_tab)
        self.name_edit = QLineEdit(self.retention_tab)
        retention_layout.addWidget(name_label)
        retention_layout.addWidget(self.name_edit)

        self.add_data_button = QPushButton("Внести данные", self.retention_tab)
        self.add_data_button.setStyleSheet("background-color: #ADFF2F; font-size: 20px;")
        QTimer.singleShot(2000, self.reset_button_text)
        retention_layout.addWidget(self.add_data_button)
        self.add_data_button.clicked.connect(self.add_data_to_excel)

        # Добавление полей на вкладку "Рейтинг ПВЗ"
        pvz_layout = QVBoxLayout(self.pvz_tab)
        pvz_layout.setSpacing(1)  # Установка расстояния между строками

        self.pvz_city_label = QLabel("Город:", self.pvz_tab)
        self.pvz_city_combo = QComboBox(self.pvz_tab)
        self.pvz_city_combo.addItems(["Выбрать город", "Екатеринбург", "Уфа", "Челябинск", "Магнитогорск"])
        self.pvz_city_combo.currentIndexChanged.connect(self.update_pvz_address_combo)
        pvz_layout.addWidget(self.pvz_city_label)
        pvz_layout.addWidget(self.pvz_city_combo)

        self.pvz_address_label = QLabel("Адрес:", self.pvz_tab)
        self.pvz_address_combo = QComboBox(self.pvz_tab)
        pvz_layout.addWidget(self.pvz_address_label)
        pvz_layout.addWidget(self.pvz_address_combo)

        self.pvz_date_label = QLabel("Дата:", self.pvz_tab)
        self.pvz_date_calendar = QCalendarWidget(self.pvz_tab)
        self.pvz_date_calendar.setFixedSize(200, 200)  # Установка размеров 200x200 пикселей
        self.pvz_date_calendar.setSelectedDate(QDate(2024, 1, 1))
        pvz_layout.addWidget(self.pvz_date_label)
        pvz_layout.addWidget(self.pvz_date_calendar)

        self.pvz_amount_label = QLabel("Сумма:", self.pvz_tab)
        self.pvz_amount_edit = QLineEdit(self.pvz_tab)
        pvz_layout.addWidget(self.pvz_amount_label)
        pvz_layout.addWidget(self.pvz_amount_edit)

        self.pvz_add_data_button = QPushButton("Внести данные", self.pvz_tab)
        self.pvz_add_data_button.setStyleSheet("background-color: #ADFF2F; font-size: 16px;")
        self.pvz_add_data_button.clicked.connect(self.add_data_to_pvz_excel)
        pvz_layout.addWidget(self.pvz_add_data_button)

        layout = QVBoxLayout(self)
        layout.addWidget(self.button)
        layout.addWidget(self.label)
        layout.addWidget(self.tab_widget)

        self.setLayout(layout)

        self.search_button.clicked.connect(self.search_data)
        self.search_reset_button.clicked.connect(self.reset_search)

    def reset_button_text(self):
        # Возвращаем исходный текст кнопки
        self.add_data_button.setText("Внести данные")

    def select_file(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "Выбрать файл", "", "Excel Files (*.xlsx *.xls)", options=options)

        if file_name:
            self.selected_file = file_name  # Сохраняем путь к выбранному файлу
            self.label.setStyleSheet("color: red; font-size: 20px;")
            self.button.setStyleSheet("background-color: orange; color: red; font-size: 20px;")
            self.button.setText("ФАЙЛ В РАБОТЕ")

            self.file_name = file_name
            self.df = pd.read_excel(file_name)
            self.show_data_in_table(self.df)

    def show_data_in_table(self, df):
        self.table_widget.setRowCount(df.shape[0])
        self.table_widget.setColumnCount(df.shape[1])
        self.table_widget.setHorizontalHeaderLabels(df.columns)

        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                item = QTableWidgetItem(str(df.iloc[i, j]))
                self.table_widget.setItem(i, j, item)

        for j in range(df.shape[1]):
            self.table_widget.setColumnWidth(j, 180)

        self.set_row_colors(df)

        self.table_widget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_widget.customContextMenuRequested.connect(self.show_context_menu)

    def set_row_colors(self, df):
        for i in range(df.shape[0]):
            status = df.iloc[i, 3]
            color = None
            if status == "К удержанию":
                color = QColor("yellow")
            elif status == "В работе":
                color = QColor("lightblue")
            elif status == "Удержано":
                color = QColor("red")
            elif status == "Доплачено":
                color = QColor("green")

            if color:
                for j in range(df.shape[1]):
                    item = self.table_widget.item(i, j)
                    if item:
                        item.setBackground(color)

    def search_data(self):
        search_text = self.search_box.text().strip().lower()
        city_filter = self.city_filter.currentText()
        if search_text or city_filter != "Выбрать город":
            self.search_reset_button.setEnabled(True)
            for i in range(self.df.shape[0]):
                found = False
                if city_filter == "Выбрать город" or city_filter.lower() in str(self.df.iloc[i, :]).lower():
                    for j in range(self.df.shape[1]):
                        if search_text in str(self.df.iloc[i, j]).lower():
                            found = True
                            continue
                if not found:
                    self.table_widget.hideRow(i)
        else:
            self.reset_search()

    def reset_search(self):
        self.search_reset_button.setEnabled(False)
        self.search_box.clear()
        self.city_filter.setCurrentIndex(0)
        for i in range(self.df.shape[0]):
            self.table_widget.showRow(i)

    def show_context_menu(self, pos):
        col = self.table_widget.currentColumn()
        if col == 3:
            menu = QMenu(self)
            menu.addAction("К удержанию", lambda: self.set_cell_status("К удержанию"))
            menu.addAction("В работе", lambda: self.set_cell_status("В работе"))
            menu.addAction("Удержано", lambda: self.set_cell_status("Удержано"))
            menu.addAction("Доплачено", lambda: self.set_cell_status("Доплачено"))
            menu.exec_(self.table_widget.mapToGlobal(pos))

    def set_cell_status(self, status):
        current_row = self.table_widget.currentRow()
        item = QTableWidgetItem(status)
        self.table_widget.setItem(current_row, 3, item)
        self.df.iloc[current_row, 3] = status  # Update the value in the DataFrame

    # Save the DataFrame to the Excel file
        self.df.to_excel(self.file_name, index=False)

    # Update row colors
        self.set_row_colors(self.df)


    def update_address_combo(self):
        city = self.city_combo.currentText()
        addresses = []

        if city == "Екатеринбург":
            addresses = ["Союзная 2", "Луначарского 182", "Пальмиро Тольятти 30", "Переулок Саперов 5"]
        elif city == "Уфа":
            addresses = ["Мустая Карима 50"]
        elif city == "Челябинск":
            addresses = ["Якутская 9", "Университетская набержная 28", "Университетская набержная 82", "Советская 65", "Сталеваров 19", "Братьев Каширенных 131а", "Энгельса 38"]
        elif city == "Магнитогорск":
            addresses = ["Грязного 15", "Советская 176"]

        self.address_combo.clear()
        self.address_combo.addItems(addresses)

     #
    def add_data_to_excel(self):
        city = self.city_combo.currentText()
        address = self.address_combo.currentText()
        barcode = self.barcode_edit.text()
        status = self.status_combo.currentText()
        amount = self.amount_edit.text()
        date = self.date_calendar.selectedDate().toString(Qt.ISODate)
        name = self.name_edit.text()
    
        if not (city and address and barcode and status and amount and name):
            self.add_data_button.setText("Заполните все поля!")
            QTimer.singleShot(2000, self.reset_button_text)
            return
    
        data = {
            "Город": [city],
            "Адрес": [address],
            "ШК": [barcode],
            "Статус удержания": [status],
            "Сумма": [amount],
            "Дата": [date],
            "ФИО": [name]
        }
    
        new_data_df = pd.DataFrame(data)
    
        if self.df is None:
            self.df = new_data_df
        else:
            self.df = pd.concat([self.df, new_data_df], ignore_index=True)
    
        self.show_data_in_table(self.df)
    
        try:
            workbook = load_workbook(self.file_name, keep_vba=True)  # Set keep_vba=True to preserve VBA code
            if "База и поиск" not in workbook.sheetnames:
                workbook.create_sheet("База и поиск")
            worksheet = workbook["База и поиск"]
    
            # Write header if the sheet is new
            if worksheet.max_row == 1:
                header = self.df.columns.tolist()
                for col, value in enumerate(header, 1):
                    worksheet.cell(row=1, column=col, value=value)
    
            for row_index, row in self.df.iterrows():
                for col_index, value in enumerate(row, 1):
                    worksheet.cell(row=row_index + 2, column=col_index, value=value)
    
            workbook.save(self.file_name)
    
            self.city_combo.setCurrentIndex(0)
            self.address_combo.clear()
            self.barcode_edit.clear()
            self.status_combo.setCurrentIndex(0)
            self.amount_edit.clear()
            self.name_edit.clear()
    
            self.add_data_button.setText("Данные внесены!")
        except Exception as e:
            print("Ошибка при сохранении файла:", e)

    def update_pvz_address_combo(self):
        city = self.pvz_city_combo.currentText()
        addresses = []

        if city == "Екатеринбург":
            addresses = ["Союзная 2", "Луначарского 182", "Пальмиро Тольятти 30", "Переулок Саперов 5"]
        elif city == "Уфа":
            addresses = ["Мустая Карима 50"]
        elif city == "Челябинск":
            addresses = ["Якутская 9", "Университетская набержная 28", "Университетская набержная 82", "Советская 65", "Сталеваров 19", "Братьев Каширенных 131а", "Энгельса 38"]
        elif city == "Магнитогорск":
            addresses = ["Грязного 15", "Советская 176"]

        self.pvz_address_combo.clear()
        self.pvz_address_combo.addItems(addresses)

    def add_data_to_pvz_excel(self):
        if not self.selected_file:
            self.label.setText("Сначала выберите файл!")
            return
    
        city = self.pvz_city_combo.currentText()
        address = self.pvz_address_combo.currentText()
        amount = self.pvz_amount_edit.text()
        date = self.pvz_date_calendar.selectedDate().toString(Qt.ISODate)
    
        if not (city and address and amount):
            self.pvz_add_data_button.setText("Заполните все поля!")
            QTimer.singleShot(2000, lambda: self.pvz_add_data_button.setText("Внести данные"))
            return
    
        data = {
            "Город": [city],
            "Адрес": [address],
            "Сумма": [amount],
            "Дата": [date]
        }
    
        pvz_df = pd.DataFrame(data)
    
        try:
            workbook = load_workbook(self.selected_file)
            if "Рейтинг ПВЗ" not in workbook.sheetnames:
                workbook.create_sheet("Рейтинг ПВЗ")
            worksheet = workbook["Рейтинг ПВЗ"]
            
            # Write header if the sheet is new
            if worksheet.max_row == 1:
                header = pvz_df.columns.tolist()
                for col, value in enumerate(header, 1):
                    worksheet.cell(row=1, column=col, value=value)
    
            for row_index, row in pvz_df.iterrows():
                for col_index, value in enumerate(row, 1):
                    worksheet.cell(row=row_index + 2, column=col_index, value=value)
    
            workbook.save(self.selected_file)
            self.pvz_city_combo.setCurrentIndex(0)
            self.pvz_address_combo.clear()
            self.pvz_amount_edit.clear()
    
            self.pvz_add_data_button.setText("Данные внесены!")
        except Exception as e:
            print("Ошибка при сохранении файла:", e)
    
    


app = QApplication(sys.argv)
window = MainWindow()
window.show()
sys.exit(app.exec_())
