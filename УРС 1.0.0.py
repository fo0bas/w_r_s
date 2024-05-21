import sys
from PyQt5.QtCore import QTimer, QDate, Qt
from PyQt5.QtGui import QKeyEvent, QColor
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import QHBoxLayout
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton, QTabWidget, 
    QFileDialog, QLabel, QLineEdit, QComboBox, QCalendarWidget, 
    QTableWidget, QTableWidgetItem, QMenu, QGridLayout, QStyle,
)
from openpyxl import Workbook, load_workbook #ПОЧЕМУ ВОРК НЕ ИНЦЛЗ Я ХЗ НО РАБОТАТ 

app = QApplication(sys.argv)
file_path = None  # ГЛОБАЛЬНАЯ ПЕРЕМЕННАЯ БИЧ

class DataEntryWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout()
        layout.setSpacing(6)
        self.city_label = QLabel("Город:")
        self.city_combo = QComboBox()
        self.city_combo.addItems(["Екатеринбург", "Уфа", "Челябинск", "Магнитогорск"])
        self.address_label = QLabel("Адрес:")
        self.address_combo = QComboBox()
        self.shk_label = QLabel("ШК:")
        self.shk_input = QLineEdit()
        self.status_label = QLabel("Статус удержания:")
        self.status_combo = QComboBox()
        self.status_combo.addItems(["К удержанию", "В работе", "Удержано", "Доплачено"])
        self.amount_label = QLabel("Сумма:")
        self.amount_input = QLineEdit()
        self.date_label = QLabel("Дата:")
        self.date_calendar = QCalendarWidget()
        self.date_calendar.setFixedSize(200, 200)
        self.fio_label = QLabel("ФИО:")
        self.fio_input = QLineEdit()
        self.add_data_button = QPushButton("Внести данные")
        self.add_data_button.setStyleSheet("background-color: green; color: white; font-weight: bold; font-size: 18px;")
        self.add_data_button.clicked.connect(self.add_data_to_excel)
        layout.addWidget(self.city_label)
        layout.addWidget(self.city_combo)
        layout.addWidget(self.address_label)
        layout.addWidget(self.address_combo)
        layout.addWidget(self.shk_label)
        layout.addWidget(self.shk_input)
        layout.addWidget(self.status_label)
        layout.addWidget(self.status_combo)
        layout.addWidget(self.amount_label)
        layout.addWidget(self.amount_input)
        layout.addWidget(self.date_label)
        layout.addWidget(self.date_calendar)
        layout.addWidget(self.fio_label)
        layout.addWidget(self.fio_input)
        layout.addWidget(self.add_data_button)
        self.setLayout(layout)

        self.update_address_combo()
        self.city_combo.currentTextChanged.connect(self.update_address_combo)

    def update_address_combo(self):
        city = self.city_combo.currentText()
        addresses = []

        if city == "Екатеринбург":
            addresses = ["Союзная 2", "Луначарского 182", "Пальмиро Тольятти 30", "Переулок Саперов 5"]
        elif city == "Уфа":
            addresses = ["Мустая Карима 50"]
        elif city == "Челябинск":
            addresses = ["Якутская 9", "Университетская набержная 28", "Университетская набержная 82",
                         "Советская 65", "Сталеваров 19", "Братьев Каширенных 131а", "Энгельса 38"]
        elif city == "Магнитогорск":
            addresses = ["Грязного 15", "Советская 176"]

        self.address_combo.clear()
        self.address_combo.addItems(addresses)

    def keyPressEvent(self, event: QKeyEvent):
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:
            self.add_data_to_excel()

    def add_data_to_excel(self):
        global file_path
        if file_path:
            amount = self.amount_input.text()
            if not amount.isdigit():
                QMessageBox.warning(self, "Ошибка", "Сумма должна быть числом.")
                return
            city = self.city_combo.currentText()
            address = self.address_combo.currentText()
            shk = self.shk_input.text()
            status = self.status_combo.currentText()
            date = self.date_calendar.selectedDate().toString("dd.MM.yyyy")
            fio = self.fio_input.text()

            if amount:  
                workbook = load_workbook(file_path)
                worksheet = workbook["Удержание"]  
                next_row = worksheet.max_row + 1

                worksheet.cell(row=next_row, column=1, value=city)
                worksheet.cell(row=next_row, column=2, value=address)
                worksheet.cell(row=next_row, column=3, value=shk)
                worksheet.cell(row=next_row, column=4, value=status)
                worksheet.cell(row=next_row, column=5, value=amount)
                worksheet.cell(row=next_row, column=6, value=fio)
                worksheet.cell(row=next_row, column=7, value=date)
                workbook.save(file_path)
                print(f"Добавлена запись в лист \"Удержание\" в Excel-файле")

                self.shk_input.clear()
                self.status_combo.setCurrentIndex(0)
                self.amount_input.clear()
                self.date_calendar.setSelectedDate(QDate.currentDate())
                self.fio_input.clear()

                self.add_data_button.setText("Успешно")
                self.add_data_button.setStyleSheet("background-color: yellow; color: black; font-weight: bold; font-size: 18px;")
                QTimer.singleShot(3000, self.reset_button_style)
                load_data_to_base_search_table()  # Обновить таблицу после внесения
            else:
                print("Необходимо ввести значение суммы.")
        else:
            QMessageBox.warning(self, "Ошибка", "Выберите файл для сохранения данных.")

    def reset_button_style(self):
        self.add_data_button.setText("Внести данные")
        self.add_data_button.setStyleSheet("background-color: green; color: white; font-weight: bold; font-size: 18px;")

def load_data_to_base_search_table():
    # Очищаем содержимое таблицы
    base_search_table.clearContents()
    workbook = load_workbook(file_path)
    worksheet = workbook["Удержание"]
    rows = worksheet.max_row
    cols = worksheet.max_column
    base_search_table.setRowCount(rows)
    base_search_table.setColumnCount(cols)
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            cell_value = worksheet.cell(row=row, column=col).value
            item = QTableWidgetItem(str(cell_value))
            base_search_table.setItem(row - 1, col - 1, item)

    # Установка имен столбцов
    column_headers = ["Город", "Адрес", "ШК", "Статус удержания", "Сумма", "ФИО", "Дата"]
    base_search_table.setHorizontalHeaderLabels(column_headers)

    # Установка цвета строки в зависимости от значения в 4 столбце
    for row in range(1, rows + 1):
        status_item = base_search_table.item(row - 1, 3)
        if status_item:
            status = status_item.text()
            color = None
            if status == "К удержанию":
                color = QColor('#FFD700')
            elif status == "В работе":
                color = QColor('#00BFFF')
            elif status == "Удержано":
                color = QColor('#FF3030')
            elif status == "Доплачено":
                color = QColor('#00C957')

            if color:
                for col in range(cols):
                    base_search_table.item(row - 1, col).setBackground(color)
    # Установка ширины столбцов
    for col in range(cols):
        base_search_table.setColumnWidth(col, 180)

    # Подключение контекстного меню к таблице
    base_search_table.setContextMenuPolicy(Qt.CustomContextMenu)
    base_search_table.customContextMenuRequested.connect(show_context_menu)


def show_context_menu(pos):#КРЧ ЭТО ВЫЗОВ ПРОСТИТУТОК ,НЕТ ВЫЗОВ МЕНЮ В ТАБЛИЦЕ УДЕРЖАНИЕ В РАБОТЕ УДЕРЖАННО ДОПЛАЧЕНО К
    col = base_search_table.currentColumn()
    if col == 3:
        menu = QMenu()
        menu.addAction("К удержанию", lambda: set_cell_status("К удержанию"))
        menu.addAction("В работе", lambda: set_cell_status("В работе"))
        menu.addAction("Удержано", lambda: set_cell_status("Удержано"))
        menu.addAction("Доплачено", lambda: set_cell_status("Доплачено"))
        menu.exec_(base_search_table.mapToGlobal(pos))

def set_cell_status(status):
    current_row = base_search_table.currentRow()
    item = QTableWidgetItem(status)
    base_search_table.setItem(current_row, 3, item)

    # обновление цветв в таблице
    set_row_colors()

    # Обновление данных в  эксельф файле что выбран 
    workbook = load_workbook(file_path)
    worksheet = workbook["Удержание"]
    worksheet.cell(row=current_row + 1, column=4, value=status)
    workbook.save(file_path)

def set_row_colors():
    rows = base_search_table.rowCount()
    cols = base_search_table.columnCount()
    for row in range(rows):
        status_item = base_search_table.item(row, 3)
        if status_item:
            status = status_item.text()
            color = None
            if status == "К удержанию":
                color = QColor('#FFD700')
            elif status == "В работе":
                color = QColor('#00BFFF')
            elif status == "#FF3030":
                color = QColor('#FF3030')
            elif status == "Доплачено":
                color = QColor('#00C957')

            if color:
                for col in range(cols):
                    base_search_table.item(row, col).setBackground(color)
####################################################
####################################################
####################################################
################РЕЙТИНГ ПВЗ ########################
class RatingPVZTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QGridLayout()

        self.city_label = QLabel("Город:")
        self.city_combo = QComboBox()
        self.city_combo.addItems(["Челябинск", "Екатеринбург", "Уфа", "Магнитогорск"])
        self.city_combo.currentTextChanged.connect(self.update_address_combo)

        self.address_label = QLabel("Адрес:")
        self.address_combo = QComboBox()

        self.deadline_label = QLabel("Выберите дату:")
        self.calendar = QCalendarWidget()
        self.calendar.setFixedSize(200,200)#РАЗМЕР КАЛЕНДАРЯ МОЖНО ПОДОГНАТЬ КРЧ 

        self.amount_label = QLabel("Сумма:")
        self.amount_input = QLineEdit()
       #КНОПКА ВНЕСТИ ДАННЫЕ ЕЕ РАЗМЕР И ЦВЕТ И ИТД 
        self.submit_button = QPushButton("Внести данные")
        self.submit_button.clicked.connect(self.on_submit)
        self.submit_button.setStyleSheet("background-color: green; color: white; font-weight: bold; font-size: 18px;")

        self.completed_label = QLabel("")
        
#крч это основной порядок строк во вкладках лушче не трогай артем 
        layout.addWidget(self.city_label, 0, 0, 1, 2)
        layout.addWidget(self.city_combo, 1, 0, 1, 2)
        layout.addWidget(self.address_label, 2, 0, 1, 2)
        layout.addWidget(self.address_combo, 3, 0, 1, 2)
        layout.addWidget(self.deadline_label, 4, 0, 1, 2)
        layout.addWidget(self.calendar, 5, 0, 1, 2)
        layout.addWidget(self.amount_label, 6, 0, 1, 2)
        layout.addWidget(self.amount_input, 7, 0, 1, 2)
        layout.addWidget(self.submit_button, 8, 0, 1, 2)
        layout.addWidget(self.completed_label, 9, 0, 1, 2)

        self.setLayout(layout)
        self.update_address_combo()

    def update_address_combo(self):
        selected_city = self.city_combo.currentText()
        addresses = []
        if selected_city == "Челябинск":
            addresses = ["Якутская 9", "Университетская набержная 28", "Университетская набержная 82",
                         "Советская 65", "Сталеваров 19", "Братьев Каширенных 131а", "Энгельса 38"]
        elif selected_city == "Екатеринбург":
            addresses = ["Союзная 2", "Луначарского 182", "Пальмиро Тольятти 30", "Переулок Саперов 5"]
        elif selected_city == "Уфа":
            addresses = ["Мустая Карима 50"]
        elif selected_city == "Магнитогорск":
            addresses = ["Грязного 15", "Советская 176"]
        self.address_combo.clear()
        self.address_combo.addItems(addresses)

  


    def on_submit(self):
        city = self.city_combo.currentText()
        address = self.address_combo.currentText()
        amount = self.amount_input.text()
        deadline = self.calendar.selectedDate().toString("dd.MM.yyyy")

        new_row = {
            "Город": city,
            "Адрес": address,
            "Сумма": amount,
            "Выбранная дата": deadline
        }

        if file_path:
            try:
                workbook = load_workbook(file_path)
                ws = workbook["Рейтинг ПВЗ"] if "Рейтинг ПВЗ" in workbook.sheetnames else workbook.create_sheet("Рейтинг ПВЗ")
            except FileNotFoundError:
                print("Ошибка: Файл не найден.")
                return

            next_row = ws.max_row + 1
            for col_idx, key in enumerate(new_row.keys(), start=1):
                ws.cell(row=next_row, column=col_idx, value=new_row[key])
            workbook.save(file_path)
            print("Данные успешно добавлены в файл:", file_path)

            

            # Очисткап поле сумма
            self.amount_input.clear()

            # Изменение стиля и текста кнопки
            self.submit_button.setText("Успешно")
            self.submit_button.setStyleSheet("background-color: yellow; color: black; font-weight: bold; font-size: 18px;")
            QTimer.singleShot(3000, self.reset_button_style)
        else:
            print("Выберите файл для сохранения данных.")
            self.completed_label.setText("Ошибка")

    def reset_button_style(self):
        self.submit_button.setText("Внести данные")
        self.submit_button.setStyleSheet("background-color: green; color: white; font-weight: bold; font-size: 18px;")

    def keyPressEvent(self, event: QKeyEvent):
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:
            self.on_submit()

##################################
#################################
############СКОРОСЬ ПРИЕМА#######
class SpeedEntryWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)

        layout = QVBoxLayout()
        layout.setSpacing(6)

        self.city_label_speed = QLabel("Город:")
        self.city_combo_speed = QComboBox()
        self.city_combo_speed.addItems(["Челябинск", "Екатеринбург", "Уфа", "Магнитогорск"])
        self.city_combo_speed.currentTextChanged.connect(self.update_address_options)

        self.address_label_speed = QLabel("Адрес:")
        self.address_combo_speed = QComboBox()

        self.deadline_label_speed = QLabel("Выберите дату:")
        self.cal_speed = QCalendarWidget()
        self.cal_speed.setFixedSize(200, 200)

        self.name_label_speed = QLabel("ФИО:")
        self.name_entry_speed = QLineEdit()

        self.amount_label_speed = QLabel("Сумма:")
        self.amount_entry_speed = QLineEdit()

        self.submit_button_speed = QPushButton("Внести данные")
        self.submit_button_speed.setStyleSheet("background-color: green; color: white; font-size: 18px;")
        self.submit_button_speed.clicked.connect(self.on_submit_speed)

        self.completed_label_speed = QLabel("")
        self.name_entry_speed.returnPressed.connect(self.on_submit_speed)
        self.amount_entry_speed.returnPressed.connect(self.on_submit_speed)

        layout.addWidget(self.city_label_speed)
        layout.addWidget(self.city_combo_speed)
        layout.addWidget(self.address_label_speed)
        layout.addWidget(self.address_combo_speed)
        layout.addWidget(self.deadline_label_speed)
        layout.addWidget(self.cal_speed)
        layout.addWidget(self.name_label_speed)
        layout.addWidget(self.name_entry_speed)
        layout.addWidget(self.amount_label_speed)
        layout.addWidget(self.amount_entry_speed)
        layout.addWidget(self.submit_button_speed)
        layout.addWidget(self.completed_label_speed)

        self.setLayout(layout)
        self.update_address_options()

    def update_address_options(self):
        selected_city = self.city_combo_speed.currentText()
        if selected_city == "Челябинск":
            addresses = ["Якутская 9", "Университетская набержная 28", "Университетская набержная 82",
                         "Советская 65", "Сталеваров 19", "Братьев Каширенных 131а", "Энгельса 38"]
        elif selected_city == "Екатеринбург":
            addresses = ["Союзная 2", "Луначарского 182", "Пальмиро Тольятти 30", "Переулок Саперов 5"]
        elif selected_city == "Уфа":
            addresses = ["Мустая Карима 50"]
        elif selected_city == "Магнитогорск":
            addresses = ["Грязного 15", "Советская 176"]
        else:
            addresses = []

        self.address_combo_speed.clear()
        self.address_combo_speed.addItems(addresses)

    def on_submit_speed(self):
        city = self.city_combo_speed.currentText()
        address = self.address_combo_speed.currentText()
        amount = self.amount_entry_speed.text()
        deadline = self.cal_speed.selectedDate().toString("dd.MM.yyyy")
        name = self.name_entry_speed.text()

        new_row = {
            "Город": city,
            "Адрес": address,
            "Сумма": amount,
            "Выбранная дата": deadline,
            "ФИО": name
        }

        global file_path
        if file_path:
            try:
                workbook = load_workbook(file_path)
                if "Скорость приема" not in workbook.sheetnames:
                    workbook.create_sheet("Скорость приема")
                worksheet = workbook["Скорость приема"]
            except FileNotFoundError:
                print("Ошибка: Файл не найден.")
                return

            next_row = worksheet.max_row + 1
            for col_idx, key in enumerate(new_row.keys(), start=1):
                worksheet.cell(row=next_row, column=col_idx, value=new_row[key])

            workbook.save(file_path)
            print("Данные успешно добавлены в файл:", file_path)

            
            self.submit_button_speed.setStyleSheet("background-color: yellow; color: black; font-size: 18px;")
            self.submit_button_speed.setText("Успешно")
            QTimer.singleShot(3000, lambda: self.completed_label_speed.setText(""))
            QTimer.singleShot(3000, lambda: self.submit_button_speed.setText("Внести данные"))
            QTimer.singleShot(3000, lambda: self.submit_button_speed.setStyleSheet("background-color: green; color: white; font-size: 18px;"))
            # Очистка полей "ФИО" и "Сумма"
            self.name_entry_speed.clear()
            self.amount_entry_speed.clear()
        else:
            print("Выберите файл для сохранения данных.")
            self.completed_label_speed.setText("Ошибка")




########ОСНОВНОЕ ОКНО КРЧ
#class MainWindow(QWidget):



class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Удержание Рейтинг Скорость version 1.0.0 @f0bas")
        self.setGeometry(100, 100, 800, 600)
        layout = QVBoxLayout()

        self.select_file_button = QPushButton("Выбрать файл") #Кнопка выбрать файл
        font = QFont("Arial", 16)  # Установка шрифта Arial размером 16
        self.select_file_button.setFont(font)
        self.select_file_button.clicked.connect(self.select_file)

        layout.addWidget(self.select_file_button)

        self.tabs = QTabWidget()
        self.tab1 = DataEntryWidget()
        self.tab2 = QWidget()
        self.tab3 = RatingPVZTab()
        self.tab4 = SpeedEntryWidget()
        font = QFont("Arial", 11)  # Выберите нужный шрифт и размер
        self.tabs.setFont(font)
        self.tabs.addTab(self.tab1, "   Удержание   ")
        self.tabs.addTab(self.tab2, "   База и поиск   ")
        self.tabs.addTab(self.tab3, "   Рейтинг ПВЗ   ")
        self.tabs.addTab(self.tab4, "   Скорость приема   ")

        layout.addWidget(self.tabs)

        self.setLayout(layout)

        self.create_tab2_ui()

    def create_tab2_ui(self):
        global base_search_table
        layout = QVBoxLayout()

        # Создание горизонтального слоя
        search_layout = QHBoxLayout()

        # Поле поиска
        self.search_line_edit = QLineEdit()
        self.search_line_edit.setPlaceholderText("Поиск")
        self.search_line_edit.setFont(QFont("Arial", 12))
        self.search_line_edit.textChanged.connect(self.filter_table)

        # Кнопка "Найти"
        self.find_button = QPushButton("Найти")
        self.find_button.setFont(QFont("Arial", 12))
        self.find_button.clicked.connect(self.find_in_table)

        # Добавление поля поиска и кнопки "Найти" в горизонтальный слой
        search_layout.addWidget(self.search_line_edit)
        search_layout.addWidget(self.find_button)

        # Выпадающий список с городами
        self.city_combo_box = QComboBox()
        self.city_combo_box.addItems(["Все города", "Екатеринбург", "Уфа", "Челябинск", "Магнитогорск"])
        self.city_combo_box.setFont(QFont("Arial", 12))
        self.city_combo_box.currentTextChanged.connect(self.filter_table)

        # Кнопка "Сбросить и обновить"
        self.reset_button = QPushButton("Сбросить и обновить")
        self.reset_button.setFont(QFont("Arial", 12))
        self.reset_button.clicked.connect(self.reset_and_refresh)

        base_search_table = QTableWidget()

        # Добавление горизонтального слоя, выпадающего списка и кнопки "Сбросить и обновить" в вертикальный слой
        layout.addLayout(search_layout)
        layout.addWidget(self.city_combo_box)
        layout.addWidget(self.reset_button)
        layout.addWidget(base_search_table)

        self.tab2.setLayout(layout)

    def select_file(self):
        global file_path
        file_path, _ = QFileDialog.getOpenFileName(self, "Выбрать файл", "", "Excel Files (*.xlsx)")
        if file_path:
            self.select_file_button.setText(f"Выбран файл: {file_path}")
            self.check_and_create_worksheets()
            load_data_to_base_search_table()
        

    def check_and_create_worksheets(self):
        global file_path
        workbook = load_workbook(file_path)
        sheet_names = workbook.sheetnames
        if "Удержание" not in sheet_names:
            workbook.create_sheet("Удержание")
        if "Рейтинг ПВЗ" not in sheet_names:
            workbook.create_sheet("Рейтинг ПВЗ")
        if "Скорость приема" not in sheet_names:
            workbook.create_sheet("Скорость приема")
            workbook.save(file_path)

    def filter_table(self):
        global base_search_table
        search_text = self.search_line_edit.text().strip()
        selected_city = self.city_combo_box.currentText()

        if not search_text and selected_city == "Все города":
            self.reset_and_refresh()
            return

        rows = base_search_table.rowCount()
        cols = base_search_table.columnCount()

        for row in range(rows):
            hide_row = True
            for col in range(cols):
                item = base_search_table.item(row, col)
                if item and (search_text.lower() in item.text().lower()) and (selected_city == "Все города" or selected_city == base_search_table.item(row, 0).text()):
                    hide_row = False
                    break
            base_search_table.setRowHidden(row, hide_row)

    def find_in_table(self):
        search_text = self.search_line_edit.text().strip()
        if search_text:
            self.filter_table()

    def reset_and_refresh(self):
        self.search_line_edit.clear()
        self.city_combo_box.setCurrentIndex(0)
        rows = base_search_table.rowCount()
        for row in range(rows):
            base_search_table.setRowHidden(row, False)


if __name__ == "__main__":
    window = MainWindow()
    window.show()
    window.showMaximized()  # Отображаем окно на максимальном разрешении экрана
    sys.exit(app.exec_())
